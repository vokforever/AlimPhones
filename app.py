import os
import re
import openpyxl
from datetime import datetime
from flask import Flask, render_template, request, send_file, flash, redirect, url_for
from werkzeug.utils import secure_filename
import tempfile
import shutil

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # В продакшене используйте безопасный ключ

# Настройки загрузки файлов
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Создаем папку для загрузок если её нет
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # Максимум 16MB

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_excel_file(filepath):
    """Обрабатывает Excel файл и возвращает найденные номера"""
    all_found_numbers = set()
    # Ищем номера с любым 3-значным кодом после +7
    phone_regex = re.compile(r'(?:[+7\s-]?\(?|8\s-?)?(\d{3})\)?[-\s]?(\d{3})[-\s]?(\d{2})[-\s]?(\d{2})')
    
    try:
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        for sheet in workbook.sheetnames:
            for row in workbook[sheet].iter_rows():
                for cell in row:
                    if cell.value:
                        matches = phone_regex.finditer(str(cell.value))
                        for match in matches:
                            # Проверяем, что код начинается с 7 (для +7)
                            code = match.group(1)
                            if code.startswith('7'):
                                normalized = f"7{code}{match.group(2)}{match.group(3)}{match.group(4)}"
                                all_found_numbers.add(normalized)
        return sorted(list(all_found_numbers))
    except Exception as e:
        raise Exception(f"Ошибка чтения файла: {e}")

def create_result_file(numbers):
    """Создает Excel файл с результатами"""
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    output_filename = f"found_numbers_+7_{timestamp}.xlsx"
    
    # Создаем временный файл
    temp_dir = tempfile.mkdtemp()
    output_filepath = os.path.join(temp_dir, output_filename)
    
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Найденные номера"
        sheet['A1'] = "Найденные номера (+7xxxxxxxxx)"
        
        for index, number in enumerate(numbers, start=2):
            sheet[f'A{index}'] = number
        
        workbook.save(output_filepath)
        return output_filepath, output_filename
    except Exception as e:
        shutil.rmtree(temp_dir)
        raise Exception(f"Ошибка создания файла результатов: {e}")

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        # Проверяем, есть ли файл в запросе
        if 'file' not in request.files:
            flash('Файл не выбран')
            return redirect(request.url)
        
        file = request.files['file']
        
        # Проверяем, выбран ли файл
        if file.filename == '':
            flash('Файл не выбран')
            return redirect(request.url)
        
        # Проверяем расширение файла
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            try:
                # Сохраняем загруженный файл
                file.save(filepath)
                
                # Обрабатываем файл
                numbers = process_excel_file(filepath)
                
                if not numbers:
                    flash('В файле не найдено номеров с кодом +7')
                    # Удаляем загруженный файл
                    os.remove(filepath)
                    return redirect(request.url)
                
                # Создаем файл результатов
                result_filepath, result_filename = create_result_file(numbers)
                
                # Удаляем загруженный файл
                os.remove(filepath)
                
                # Возвращаем файл для скачивания
                return send_file(
                    result_filepath,
                    as_attachment=True,
                    download_name=result_filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
            except Exception as e:
                flash(f'Ошибка обработки файла: {str(e)}')
                # Удаляем загруженный файл в случае ошибки
                if os.path.exists(filepath):
                    os.remove(filepath)
                return redirect(request.url)
        else:
            flash('Разрешены только файлы Excel (.xlsx, .xls)')
            return redirect(request.url)
    
    return render_template('upload.html')

@app.route('/health')
def health():
    return {'status': 'healthy'}

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=80)
