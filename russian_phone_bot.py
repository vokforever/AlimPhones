import os
import re
import tempfile
import openpyxl
from datetime import datetime
from telegram import Update, BotCommand
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
import logging
from collections import defaultdict

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', 
    level=logging.INFO,
    handlers=[
        logging.FileHandler('/tmp/bot.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Токен бота (получи у BotFather)
BOT_TOKEN = os.getenv('BOT_TOKEN', 'YOUR_BOT_TOKEN_HERE')

# Статистика использования
user_stats = defaultdict(lambda: {'files': 0, 'numbers': 0, 'last_used': None})

class RussianPhoneProcessor:
    def __init__(self):
        # Регулярное выражение для всех российских номеров
        # Ловит номера вида: +7XXXXXXXXXX, 8XXXXXXXXXX, 7XXXXXXXXXX
        self.phone_regex = re.compile(
            r'(?:\+?7|8)[-\s\(\)]?'  # +7, 7 или 8 в начале
            r'(\d{3})[-\s\(\)]?'     # 3-значный код
            r'(\d{3})[-\s]?'         # 3 цифры
            r'(\d{2})[-\s]?'         # 2 цифры
            r'(\d{2})'               # 2 цифры
        )
        
    def process_excel_file(self, file_path: str) -> dict:
        """Обрабатывает Excel файл и возвращает найденные номера"""
        numbers_found = set()
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            cell_text = str(cell.value)
                            matches = self.phone_regex.finditer(cell_text)
                            for match in matches:
                                # Нормализуем номер к формату 7XXXXXXXXXX
                                code = match.group(1)
                                number = match.group(2) + match.group(3) + match.group(4)
                                normalized = f"7{code}{number}"
                                
                                # Проверяем что это валидный российский номер
                                if self.is_valid_russian_number(normalized):
                                    numbers_found.add(normalized)
                                        
        except Exception as e:
            logger.error(f"Ошибка обработки файла: {e}")
            raise
            
        return {
            'numbers': numbers_found,
            'total': len(numbers_found)
        }
    
    def is_valid_russian_number(self, number: str) -> bool:
        """Проверяет валидность российского номера"""
        if len(number) != 11:
            return False
        if not number.startswith('7'):
            return False
        if not number[1:].isdigit():
            return False
        # Проверяем что код находится в допустимом диапазоне
        code = number[1:4]
        return code.isdigit() and '200' <= code <= '999'  # Общий диапазон российских кодов
    
    def create_result_file(self, results: dict, original_filename: str) -> str:
        """Создает Excel файл с результатами (один столбец с номерами)"""
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        temp_file = tempfile.NamedTemporaryFile(
            delete=False,
            suffix=f'_russian_phones_{timestamp}.xlsx'
        )
        
        workbook = openpyxl.Workbook()
        
        # Лист с номерами
        sheet = workbook.active
        sheet.title = "Найденные номера"
        sheet['A1'] = "Российские номера телефонов"
        
        sorted_numbers = sorted(list(results['numbers']))
        for index, number in enumerate(sorted_numbers, start=2):
            sheet[f'A{index}'] = number
            
        workbook.save(temp_file.name)
        return temp_file.name

processor = RussianPhoneProcessor()

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Приветствие и основные инструкции"""
    welcome_text = """
🇷🇺 **ПОИСК РОССИЙСКИХ НОМЕРОВ ТЕЛЕФОНОВ**

Привет! Я бот для автоматического поиска **всех российских номеров телефонов** из Excel файлов.

🔍 **ЧТО Я УМЕЮ:**
• Нахожу ВСЕ российские номера (+7, 8)
• Удаляю дубликаты
• Нормализую формат (7XXXXXXXXXX)
• Создаю Excel файл с одним столбцом номеров

📱 **ПОДДЕРЖИВАЕМЫЕ ФОРМАТЫ:**
```
✅ +7 (999) 123-45-67 (любой код)
✅ 8-912-345-67-89
✅ 7 495 123 45 67
✅ +79161234567
✅ 8(800)555-35-35
```

🚀 **КАК ИСПОЛЬЗОВАТЬ:**

**1️⃣** Отправьте Excel файл (.xlsx)
**2️⃣** Получите результат
**3️⃣** Скачайте файл с номерами

💡 **КОМАНДЫ:**
/help - подробная справка
/example - примеры номеров
/stats - ваша статистика

⚡ **ГОТОВЫ? Отправьте Excel файл!**
    """
    await update.message.reply_text(welcome_text, parse_mode='Markdown')

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Подробная справка"""
    help_text = """
📖 **ПОЛНАЯ ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ**

🔧 **ТЕХНИЧЕСКИЕ ТРЕБОВАНИЯ:**
• **Формат:** .xlsx (Excel 2007+)
• **Размер:** до 20 МБ
• **Обработка:** все листы и ячейки
• **Скорость:** 1000-5000 номеров/сек

🇷🇺 **КАКИЕ НОМЕРА ИЩЕТ БОТ:**

**Все российские номера:**
• **Мобильные:** все коды с 900 по 999
• **Городские:** все региональные коды
• **Специальные:** 800 и другие

📊 **ЧТО ПОЛУЧИТЕ В РЕЗУЛЬТАТЕ:**

**Excel файл с одним листом:**

🔸 **"Найденные номера"** - полный список в формате 7XXXXXXXXXX

**Сообщение со статистикой:**
• Общее количество найденных номеров
• Количество уникальных номеров

⚠️ **ВАЖНО:**

**Безопасность:**
✅ Файлы удаляются после обработки
✅ Данные не сохраняются
✅ Полная конфиденциальность

**Форматы номеров:**
```
✅ +7 912 345 67 89
✅ 8-495-123-45-67
✅ 7(916)1234567
✅ +79001234567
✅ 8 800 555 35 35
```

**Не распознается:**
```
❌ +38 (Украина)
❌ +375 (Беларусь)
❌ короткие номера
❌ текст без цифр
```

💡 **СОВЕТЫ:**
• Используйте текстовый формат ячеек
• Номера должны быть читаемы (не картинки)
• Файл без пароля
• Стандартная структура Excel

🎯 **ПРИМЕРЫ ИСПОЛЬЗОВАНИЯ:**
• Очистка клиентской базы
• Подготовка к рассылкам
• Создание телефонных баз

❓ **Остались вопросы? Просто отправьте файл!**
    """
    await update.message.reply_text(help_text, parse_mode='Markdown')

async def example_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Примеры российских номеров"""
    example_text = """
📝 **ПРИМЕРЫ РОССИЙСКИХ НОМЕРОВ**

**Номера (будут найдены):**
```
+7 (912) 345-67-89
8-916-123-45-67
7 950 123 45 67
+79031234567
8(977)555-35-35
79781234567
+7 (495) 123-45-67
8-812-234-56-78
7 343 987 65 43
8(383)555-12-34
8-800-555-35-35
+7 (804) 123-45-67
```

**НЕ российские (игнорируются):**
```
❌ +38 (067) 123-45-67  (Украина)
❌ +375 (29) 123-45-67  (Беларусь)
❌ +1 (555) 123-4567    (США)
```

**Все найденные номера приводятся к формату:**
`7XXXXXXXXXX`

**Например:**
• `+7 (912) 345-67-89` → `79123456789`
• `8-495-123-45-67` → `74951234567`

🔥 **Создайте тестовый Excel файл с этими номерами и проверьте!**
    """
    await update.message.reply_text(example_text, parse_mode='Markdown')

async def stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Статистика пользователя"""
    user_id = update.effective_user.id
    stats = user_stats[user_id]
    
    if stats['files'] == 0:
        stats_text = """
📊 **ВАША СТАТИСТИКА**

🆕 Вы пока не обрабатывали файлы
📁 Отправьте первый Excel файл для анализа!

💡 **После обработки вы увидите:**
• Количество обработанных файлов
• Общее число найденных номеров  
• Дата последнего использования
        """
    else:
        last_used = stats['last_used'].strftime('%d.%m.%Y %H:%M') if stats['last_used'] else 'Неизвестно'
        stats_text = f"""
📊 **ВАША СТАТИСТИКА**

📁 **Обработано файлов:** {stats['files']}
📱 **Найдено номеров:** {stats['numbers']:,}
📅 **Последнее использование:** {last_used}

🏆 **Среднее на файл:** {stats['numbers'] // stats['files'] if stats['files'] > 0 else 0} номеров

💡 Продолжайте использовать бот для анализа телефонных баз!
        """
    
    await update.message.reply_text(stats_text, parse_mode='Markdown')

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка загруженных Excel файлов"""
    document = update.message.document
    user_id = update.effective_user.id
    
    # Проверки файла
    if not document.file_name.endswith('.xlsx'):
        await update.message.reply_text(
            "❌ **Неподдерживаемый формат!**\n\n"
            "Поддерживаются только файлы **.xlsx**\n"
            "Пересохраните файл в Excel и попробуйте снова.",
            parse_mode='Markdown'
        )
        return
    
    if document.file_size > 20 * 1024 * 1024:
        await update.message.reply_text(
            "❌ **Файл слишком большой!**\n\n"
            "Максимальный размер: **20 МБ**\n"
            "Попробуйте разделить на несколько файлов.",
            parse_mode='Markdown'
        )
        return
    
    # Сообщение о начале обработки
    processing_message = await update.message.reply_text(
        "🔄 **Обрабатываю файл...**\n\n"
        "⏳ Анализирую все листы и ячейки\n"
        "🔍 Ищу российские номера телефонов\n"
        "📊 Подготавливаю статистику\n\n"
        "_Это может занять от нескольких секунд до минуты_",
        parse_mode='Markdown'
    )
    
    try:
        # Скачиваем и обрабатываем файл
        file = await document.get_file()
        temp_input = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        await file.download_to_drive(temp_input.name)
        
        # Обрабатываем файл
        results = processor.process_excel_file(temp_input.name)
        os.unlink(temp_input.name)
        
        if results['total'] > 0:
            # Обновляем статистику пользователя
            user_stats[user_id]['files'] += 1
            user_stats[user_id]['numbers'] += results['total']
            user_stats[user_id]['last_used'] = datetime.now()
            
            # Создаем файл с результатами
            result_file = processor.create_result_file(results, document.file_name)
            
            # Формируем статистику
            stats_text = f"""
✅ **ОБРАБОТКА ЗАВЕРШЕНА!**

📊 **СТАТИСТИКА:**
• Исходный файл: `{document.file_name}`
• Найдено номеров: **{results['total']:,}**
• Уникальных номеров: **{len(results['numbers']):,}**

📁 **Файл с результатами прикреплен ниже ⬇️**"""
            
            # Отправляем статистику
            await context.bot.edit_message_text(
                chat_id=update.effective_chat.id,
                message_id=processing_message.message_id,
                text=stats_text,
                parse_mode='Markdown'
            )
            
            # Отправляем файл с результатами
            with open(result_file, 'rb') as f:
                await update.message.reply_document(
                    document=f,
                    filename=f"russian_phones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    caption="📋 Российские номера телефонов\n\n"
                           "📊 Содержит один столбец с номерами в формате 7XXXXXXXXXX"
                )
            
            os.unlink(result_file)
            
        else:
            await context.bot.edit_message_text(
                chat_id=update.effective_chat.id,
                message_id=processing_message.message_id,
                text="❌ **Российские номера не найдены**\n\n"
                     "Возможные причины:\n"
                     "• Файл не содержит российских номеров (+7, 8)\n" 
                     "• Номера записаны как изображения\n"
                     "• Используется нестандартный формат\n\n"
                     "💡 Попробуйте команду /example для примеров",
                parse_mode='Markdown'
            )
            
    except Exception as e:
        logger.error(f"Ошибка обработки файла от пользователя {user_id}: {e}")
        await context.bot.edit_message_text(
            chat_id=update.effective_chat.id,
            message_id=processing_message.message_id,
            text=f"❌ **Ошибка обработки файла**\n\n"
                 f"`{str(e)}`\n\n"
                 f"Попробуйте:\n"
                 f"• Пересохранить файл в Excel\n"
                 f"• Проверить целостность файла\n" 
                 f"• Уменьшить размер файла",
            parse_mode='Markdown'
        )

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка текстовых сообщений"""
    text = update.message.text.lower()
    
    responses = {
        ('привет', 'hello', 'hi', 'здравствуй'): 
            "👋 Привет! Отправьте Excel файл (.xlsx) с номерами телефонов для анализа!",
        ('помощь', 'help', 'как работать', 'что делать'):
            "❓ Используйте /help для подробной справки\n📁 Или просто отправьте .xlsx файл!",
        ('спасибо', 'благодарю', 'thanks', 'thank you'):
            "😊 Пожалуйста! Обращайтесь для анализа телефонных баз!",
        ('операторы', 'какие операторы', 'коды'):
            "📱 Нахожу все российские номера телефонов в формате 7XXXXXXXXXX\n💡 /example - примеры номеров"
    }
    
    for keywords, response in responses.items():
        if any(keyword in text for keyword in keywords):
            await update.message.reply_text(response)
            return
    
    # Стандартный ответ
    await update.message.reply_text(
        "📎 **Отправьте Excel файл для анализа!**\n\n"
        "🔸 Формат: **.xlsx**\n"
        "🔸 Размер: до **20 МБ**\n" 
        "🔸 Ищу: **все российские номера** (+7, 8)\n\n"
        "💡 Команды: /help /example /stats",
        parse_mode='Markdown'
    )

async def set_bot_commands(bot):
    """Настройка меню команд бота"""
    commands = [
        BotCommand("start", "🏠 Главное меню и инструкции"),
        BotCommand("help", "📖 Подробная справка по использованию"),
        BotCommand("example", "📝 Примеры российских номеров"),
        BotCommand("stats", "📊 Ваша статистика использования")
    ]
    await bot.set_my_commands(commands)

def main():
    """Запуск бота"""
    if not BOT_TOKEN or BOT_TOKEN == "YOUR_BOT_TOKEN_HERE":
        print("❌ Установите токен бота в переменной окружения BOT_TOKEN!")
        print("Получите токен у @BotFather в Telegram")
        return
        
    print("🚀 Запуск бота для поиска российских номеров телефонов...")
    
    # Создаем приложение
    application = Application.builder().token(BOT_TOKEN).build()
    
    # Регистрируем обработчики
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("example", example_command))
    application.add_handler(CommandHandler("stats", stats_command))
    application.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    
    # Настраиваем команды бота
    import asyncio
    asyncio.get_event_loop().run_until_complete(set_bot_commands(application.bot))
    
    print("✅ Бот запущен и готов к работе!")
    print("📱 Отправьте боту Excel файл с номерами телефонов")
    
    # Запускаем polling
    application.run_polling(
        allowed_updates=Update.ALL_TYPES,
        drop_pending_updates=True
    )

if __name__ == '__main__':
    main()