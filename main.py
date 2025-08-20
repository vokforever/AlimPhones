import os
import re
import openpyxl
from datetime import datetime

INPUT_DIR = 'in'
OUTPUT_DIR = 'out'
BASE_OUTPUT_NAME = 'found_numbers_978'


def setup_directories():
    if not os.path.isdir(INPUT_DIR):
        print(f"Создаю папку для исходных файлов: '{INPUT_DIR}'")
        os.makedirs(INPUT_DIR)

    if not os.path.isdir(OUTPUT_DIR):
        print(f"Создаю папку для результатов: '{OUTPUT_DIR}'")
        os.makedirs(OUTPUT_DIR)


def run_processor():
    setup_directories()

    all_found_numbers = set()
    phone_regex = re.compile(r'(?:[+7\s-]?\(?|8\s-?)?(978)\)?[-\s]?(\d{3})[-\s]?(\d{2})[-\s]?(\d{2})')

    files_to_process = [f for f in os.listdir(INPUT_DIR) if f.endswith('.xlsx')]

    if not files_to_process:
        print(f"\nПапка '{INPUT_DIR}' пуста. Поместите в нее xlsx файлы и запустите скрипт снова.")
        return

    print(f"Начинаю обработку {len(files_to_process)} файлов из папки '{INPUT_DIR}'...")

    for filename in files_to_process:
        filepath = os.path.join(INPUT_DIR, filename)
        print(f"\n--- Анализирую файл: {filename} ---")

        numbers_in_this_file = set()
        try:
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            for sheet in workbook.sheetnames:
                for row in workbook[sheet].iter_rows():
                    for cell in row:
                        if cell.value:
                            matches = phone_regex.finditer(str(cell.value))
                            for match in matches:
                                normalized = f"7{match.group(1)}{match.group(2)}{match.group(3)}{match.group(4)}"
                                numbers_in_this_file.add(normalized)
        except Exception as e:
            print(f"Ошибка чтения файла {filename}: {e}")
            continue

        if numbers_in_this_file:
            print(f"Найдено {len(numbers_in_this_file)} уникальных номеров.")
            all_found_numbers.update(numbers_in_this_file)

            # --- БЛОК ИЗМЕНЕН ---
            try:
                # 1. Изменен вопрос
                answer = input(f"Удалить обработанный файл '{filename}'? (да/нет): ").lower().strip()
                if answer in ['да', 'д', 'yes', 'y']:
                    # 2. Изменено действие с os.rename на os.remove
                    os.remove(filepath)
                    # 3. Изменено сообщение
                    print(f"Файл '{filename}' удален.")
            except Exception as e:
                # 4. Изменено сообщение об ошибке
                print(f"Не удалось удалить файл. Ошибка: {e}")
            # --- КОНЕЦ ИЗМЕНЕННОГО БЛОКА ---
        else:
            print("Номера с кодом 978 в этом файле не найдены.")

    if all_found_numbers:
        save_results(sorted(list(all_found_numbers)))
    else:
        print("\nОбработка завершена. Номеров с кодом 978 не найдено ни в одном файле.")


def save_results(numbers):
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    output_filename = f"{BASE_OUTPUT_NAME}_{timestamp}.xlsx"
    output_filepath = os.path.join(OUTPUT_DIR, output_filename)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Найденные номера"
    sheet['A1'] = "Найденные номера (7978xxxxxxx)"

    for index, number in enumerate(numbers, start=2):
        sheet[f'A{index}'] = number

    try:
        workbook.save(output_filepath)
        print(f"\n========================================================")
        print(f"Готово! Всего найдено {len(numbers)} уникальных номеров.")
        print(f"Результаты сохранены в файл: {os.path.abspath(output_filepath)}")
        print(f"========================================================")
    except Exception as e:
        print(f"\nНе удалось сохранить файл результатов. Ошибка: {e}")


if __name__ == "__main__":
    run_processor()